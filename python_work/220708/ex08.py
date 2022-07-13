import datetime
from openpyxl import load_workbook

wb = load_workbook('a.xlsx')
ws = wb.create_sheet('test')

ws['A1'] = datetime.datetime(2010,7,21)
ws['A2'] = datetime.datetime.now()

print(ws['A2'].value)

wb.save('aa.xlsx')
wb.close()