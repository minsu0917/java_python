from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = '123123'
ws['A2'] = 'hello'
ws.title = '처음 자동'
ws.sheet_properties.tabColor = "1072BA"
ms1 = wb.create_sheet("Mysheet")
ms1['B1'] = 'hi hi'
ms1['B2'] = 'hahaha'

ms2 = wb.create_sheet("Mysheet",0)
ms2['C1'] = 'test'
ms2['C2'] = 'test'


wb.save('a.xlsx')
wb.close()